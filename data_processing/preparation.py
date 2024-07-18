import openpyxl
import pandas as pd
from datetime import datetime
import os
import shutil
from openpyxl import load_workbook

from data_processing.loaders import DatenLaden

class DataPreparation(DatenLaden):
    def __init__(self):
        super().__init__()  # Initialisiere die Superklasse, um Zugriff auf deren Variablen zu erhalten
        self.template_path = 'templates/SAP_TSA_Template.xlsx'
        self.workbook = None
        self.load_template()

    def load_template(self):
        # Lade das Template und erstelle eine Kopie
        self.workbook = openpyxl.load_workbook(self.template_path)
        self.workbook.save('working_copy.xlsx')  # Speichere die Kopie unter einem neuen Namen

    def enrich_dataframe(self):
        # Anreichern des DataFrames mit zusätzlichen Informationen
        if self.data is not None:
            # Stelle sicher, dass 'Date' als String vorliegt
            self.data['Date'] = self.data['Date'].astype(str)

            # Korrigiere das Datum für Oktober
            self.data['Date'] = self.data['Date'].apply(
                lambda x: x[:-1] + '10' if x.endswith('.1') else x
            )

            # Monatsnamen extrahieren
            self.data['Month'] = self.data['Date'].apply(
                lambda x: datetime.strptime(x, '%Y.%m').strftime('%b')
            )

            # Audit Period bestimmen
            audit_period_mapping = {'Q1': 3, 'Q2': 6, 'Q3': 9, 'Q4': 12}
            audit_period_months = audit_period_mapping[self.quartal]
            last_month = self.data['Date'].max()
            last_year = int(last_month.split('.')[0])
            last_month = int(last_month.split('.')[1])

            def determine_period(date):
                year, month = map(int, date.split('.'))
                if year == last_year and month > (last_month - audit_period_months):
                    return 'Audit Period'
                else:
                    return 'Prior Period'

            self.data['Period'] = self.data['Date'].apply(determine_period)

            # Fiscal Year bestimmen
            def determine_fiscal_year(date, period):
                year, month = map(int, date.split('.'))
                if period == 'Audit Period':
                    fiscal_year = last_year % 100
                else:
                    months_difference = (last_year - year) * 12 + (last_month - month)
                    fiscal_year = (last_year - (months_difference // 12)) % 100
                return f'FY {fiscal_year:02d}'

            self.data['Fiscal_Year'] = self.data.apply(lambda row: determine_fiscal_year(row['Date'], row['Period']), axis=1)

            # Sortiere den DataFrame nach dem Datum
            self.data = self.data.sort_values('Date')

    def write_to_excel_template(self):
        # Erstelle den Output-Ordner, falls er nicht existiert
        output_dir = os.path.join(os.getcwd(), 'output')
        os.makedirs(output_dir, exist_ok=True)

        # Kopiere das Template
        template_path = os.path.join('templates', 'SAP_TSA_Template.xlsx')
        output_filename = f"{self.gesellschaft}_{self.account}_{self.jahr}_TSA_Doku.xlsx"
        output_path = os.path.join(output_dir, output_filename)
        shutil.copy(template_path, output_path)

        # Lade die kopierte Arbeitsmappe
        workbook = load_workbook(output_path)
        sheet = workbook['1. Data Validation']

        # Schreibe die Daten in die Excel-Datei
        for i, row in enumerate(self.data.itertuples(), start=28):
            sheet[f'B{i}'] = row.Date
            sheet[f'C{i}'] = row.Month
            sheet[f'D{i}'] = row.Fiscal_Year
            sheet[f'E{i}'] = row.Period

        # Fülle G27 mit dem Wert aus self.account + " total"
        sheet['G27'] = f"{self.account} total"

        # Finde die Spalte, die "_total" enthält
        total_column = next((col for col in self.data.columns if '_total' in col), None)

        if total_column:
            # Fülle G28 und darunter mit den Werten aus der "_total" Spalte
            for i, value in enumerate(self.data[total_column], start=28):
                sheet[f'G{i}'] = value

        # Speichere die Änderungen
        workbook.save(output_path)
        print(f"Excel-Datei wurde erstellt: {output_path}")

        # Speichere alle Spaltenüberschriften
        self.spaltenueberschriften = self.data.columns.tolist()


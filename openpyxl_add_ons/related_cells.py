from openpyxl.utils import get_column_letter, column_index_from_string

def get_cell_name(start_cell, x_offset, y_offset):
    """
    Berechnet den Namen einer Zelle basierend auf einer Ausgangszelle und Offsets.
    
    :param start_cell: Ausgangszelle als String (z.B. 'A1')
    :param x_offset: Horizontale Verschiebung (positiv = rechts, negativ = links)
    :param y_offset: Vertikale Verschiebung (positiv = unten, negativ = oben)
    :return: Name der Zielzelle als String
    """
    col, row = column_index_from_string(start_cell[0]), int(start_cell[1:])
    
    new_col = col + x_offset
    new_row = row + y_offset
    
    if new_col < 1 or new_row < 1:
        raise ValueError("Die Verschiebung führt zu einer ungültigen Zelle.")
    
    return f"{get_column_letter(new_col)}{new_row}"

def get_cell_value(sheet, start_cell, x_offset, y_offset):
    """
    Gibt den Wert einer Zelle basierend auf einer Ausgangszelle und Offsets zurück.
    
    :param sheet: Das Arbeitsblatt-Objekt
    :param start_cell: Ausgangszelle als String (z.B. 'A1')
    :param x_offset: Horizontale Verschiebung (positiv = rechts, negativ = links)
    :param y_offset: Vertikale Verschiebung (positiv = unten, negativ = oben)
    :return: Wert der Zielzelle
    """
    try:
        target_cell = get_cell_name(start_cell, x_offset, y_offset)
        return sheet[target_cell].value
    except ValueError as e:
        print(f"Fehler: {str(e)}")
        return None

def cell_below(sheet, cell):
    row_below = cell.row + 1
    cell_below = sheet.cell(row=row_below, column=cell.column)
    return cell_below.value


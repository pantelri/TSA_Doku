from openpyxl.utils import get_column_letter

def write_basic_data(sheet, data):
    for i, row in enumerate(data.itertuples(), start=28):
        sheet[f'B{i}'] = row.Date
        sheet[f'C{i}'] = row.Month
        sheet[f'D{i}'] = row.Fiscal_Year
        sheet[f'E{i}'] = row.Period

def write_total_data(sheet, data, account):
    sheet['G27'] = f"{account} total"
    total_column = next((col for col in data.columns if '_total' in col), None)
    if total_column:
        for i, value in enumerate(data[total_column], start=28):
            sheet[f'G{i}'] = value

def write_subtotal_data(sheet, data, account, account_name):
    subtotal_columns = [col for col in data.columns if f"{account_name}_subtotal" in col]
    if len(subtotal_columns) > 3:
        raise ValueError("Doku-Template wird bisher nur für max. 3 Subtotal-Spalten des Accounts_to_audit unterstützt")
    for idx, col in enumerate(subtotal_columns):
        cell = chr(ord('H') + idx)  # H, I, J
        suffix = col.split(f"{account_name}_subtotal_")[-1]
        sheet[f'{cell}27'] = f"{account} {suffix}"
        for i, value in enumerate(data[col], start=28):
            sheet[f'{cell}{i}'] = value

def write_cos_data(sheet, data):
    cos_columns = [col for col in data.columns if col.startswith("COS")]
    if cos_columns:
        last_subtotal_col = chr(ord('H') + len([col for col in data.columns if "_subtotal" in col]) - 1)
        start_col = chr(ord(last_subtotal_col) + 1)
        
        if len(cos_columns) > 1:
            sheet.insert_cols(ord(start_col) - ord('A') + 1, len(cos_columns) - 1)
        
        for idx, col in enumerate(cos_columns):
            cell = chr(ord(start_col) + idx)
            header = col.replace('_', ' ')
            sheet[f'{cell}27'] = header
            for i, value in enumerate(data[col], start=28):
                sheet[f'{cell}{i}'] = value

def write_volume_data(sheet, data):
    volume_columns = [col for col in data.columns if 'Volume' in col]
    if len(volume_columns) > 3:
        sheet.insert_cols(15, len(volume_columns) - 3)
    for idx, col in enumerate(volume_columns):
        cell = get_column_letter(13 + idx)  # M, N, O, ...
        header = col.replace('_', ' ')
        sheet[f'{cell}27'] = header
        for i, value in enumerate(data[col], start=28):
            sheet[f'{cell}{i}'] = value

def write_index_data(sheet, data):
    index_columns = [col for col in data.columns if 'index_' in col]
    if len(index_columns) > 1:
        sheet.insert_cols(18, len(index_columns) - 1)
        for col_idx in range(18, 18 + len(index_columns) - 1):
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.font = sheet['Q' + str(cell.row)].font.copy()
                    cell.border = sheet['Q' + str(cell.row)].border.copy()
                    cell.fill = sheet['Q' + str(cell.row)].fill.copy()
                    cell.number_format = sheet['Q' + str(cell.row)].number_format
                    cell.alignment = sheet['Q' + str(cell.row)].alignment.copy()
    
    for idx, col in enumerate(index_columns):
        cell = get_column_letter(17 + idx)  # Q, R, ...
        header = f"price {col.replace('_', ' ')}"
        sheet[f'{cell}27'] = header
        for i, value in enumerate(data[col], start=28):
            sheet[f'{cell}{i}'] = value
    


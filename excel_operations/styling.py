from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

def auto_adjust_column_width(ws):
    """
    Auto adjust the width of the columns in a worksheet ('ws') based on the longest entry in each column.
    """
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length


def format_headers(ws, header_row, styles):
    """
    Format the header row of a worksheet ('ws') using a provided styles dictionary.
    The dictionary can include font name, size, bold, italic, underline, color, fill color, and alignment.
    """
    for cell in ws[header_row]:
        if 'font_name' in styles:
            cell.font = Font(name=styles['font_name'], size=styles.get('font_size', 11),
                             bold=styles.get('bold', False), italic=styles.get('italic', False),
                             underline=styles.get('underline', 'none'), color=styles.get('font_color', '000000'))
        if 'fill_color' in styles:
            cell.fill = PatternFill(start_color=styles['fill_color'], end_color=styles['fill_color'], fill_type="solid")
        if 'alignment' in styles:
            cell.alignment = Alignment(horizontal=styles.get('alignment', 'center'))

# Stil-Dictionary
header_styles = {
    'font_name': 'Calibri',
    'font_size': 12,
    'bold': True,
    'font_color': 'FFFFFF', # Weiß
    'fill_color': '000000',  # Schwarz
    'alignment': 'center'
}
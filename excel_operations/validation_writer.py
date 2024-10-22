from excel_operations.validation_writer_functions import (
    write_basic_data,
    write_total_data,
    write_subtotal_data,
    write_cos_data,
    write_volume_data,
    write_index_data,
    cell_below
)
from openpyxl.utils import get_column_letter
import statistics

class Validation():
    def __init__(self, ExcelWriter):
        self.validation_sheet = ExcelWriter.validation_sheet
        self.data = ExcelWriter.data
        self.account = ExcelWriter.account
        self.account_name = ExcelWriter.account_name

    def fill_worksheet(self):
        write_basic_data(self.validation_sheet, self.data)
        write_total_data(self.validation_sheet, self.data, self.account)
        write_subtotal_data(self.validation_sheet, self.data, self.account, self.account_name)
        write_cos_data(self.validation_sheet, self.data)
        write_volume_data(self.validation_sheet, self.data)
        write_index_data(self.validation_sheet, self.data)
        self.copy_row_27_to_row_6()
        self.calculate_and_write_summaries()
        self.remove_empty_columns()

    def copy_row_27_to_row_6(self):
        if self.validation_sheet is None:
            return

        for col in range(7, 27):  # G to Z
            col_letter = get_column_letter(col)
            self.validation_sheet[f'{col_letter}6'] = self.validation_sheet[f'{col_letter}27'].value

    def calculate_and_write_summaries(self):
        if self.validation_sheet is None:
            return

        for col in range(7, 27):  # G to Z
            col_letter = get_column_letter(col)
            cell_value = self.validation_sheet[f'{col_letter}6'].value
            
            if cell_value is None:
                continue

            is_price_index = str(cell_value).lower().startswith("price index")

            for row, range_start, range_end in [(7, 28, 39), (11, 40, 51), (15, 52, 63)]:
                values = [self.validation_sheet[f'{col_letter}{i}'].value for i in range(range_start, range_end + 1)]
                values = [v for v in values if v is not None and isinstance(v, (int, float))]
                
                if values:
                    if is_price_index:
                        result = statistics.mean(values)
                    else:
                        result = sum(values)
                    
                    self.validation_sheet[f'{col_letter}{row}'] = result

    def remove_empty_columns(self):
        if self.validation_sheet is None:
            return

        columns_to_delete = []
        last_column = None

        for cell in self.validation_sheet[27]:
            wert_darunter = cell_below(self.validation_sheet, cell)
            if cell.column_letter not in ['A', 'F'] and cell.value is None:
                columns_to_delete.append(cell.column)
            elif wert_darunter is not None:
                if isinstance(wert_darunter.value, (int, float)) and wert_darunter.value > 0:
                    last_column = cell.column
                elif isinstance(wert_darunter.value, str) and wert_darunter.value.strip():
                    last_column = cell.column

        print(last_column)
        if last_column:
            # Hide the first column after last_column
            col_to_hide = self.validation_sheet.column_dimensions[get_column_letter(last_column + 1)]
            col_to_hide.hidden = True

        for column_index in reversed(columns_to_delete):
            self.validation_sheet.delete_cols(column_index)
